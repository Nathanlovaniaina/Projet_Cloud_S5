import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Configuration
MD_FILE = "repartition_taches.md"
EXCEL_FILE = "repartition_taches.xlsx"

def parse_markdown(file_path):
    """Parse le fichier markdown et extrait les tâches"""
    tasks = []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Regex pour trouver chaque tâche
    # Cherche les patterns "## Tâche X" et les champs qui suivent
    task_pattern = r"## Tâche \d+\n\*\*Catégorie\*\* : ([^\n]+)\n\*\*Module\*\* : ([^\n]+)\n\*\*Tâches\*\* : ([^\n]+)\n\*\*Type\*\* : ([^\n]+)\n\*\*Qui\*\* : ([^\n]+)\n\*\*Estimation\*\* : ([^\n]+)"
    
    matches = re.finditer(task_pattern, content)
    
    for match in matches:
        task = {
            'categorie': match.group(1).strip(),
            'module': match.group(2).strip(),
            'taches': match.group(3).strip(),
            'type': match.group(4).strip(),
            'qui': match.group(5).strip(),
            'estimation': match.group(6).strip()
        }
        tasks.append(task)
    
    return tasks

def create_excel(tasks, output_file):
    """Crée un fichier Excel avec les tâches"""
    
    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tâches"
    
    # Définir les colonnes
    columns = ['Catégorie', 'Module', 'Tâches', 'Type', 'Qui', 'Estimation']
    
    # Ajouter l'en-tête
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        
        # Formater l'en-tête
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Ajouter les données
    for row_idx, task in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1).value = task['categorie']
        ws.cell(row=row_idx, column=2).value = task['module']
        ws.cell(row=row_idx, column=3).value = task['taches']
        ws.cell(row=row_idx, column=4).value = task['type']
        ws.cell(row=row_idx, column=5).value = task['qui']
        ws.cell(row=row_idx, column=6).value = task['estimation']
        
        # Formater les cellules
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 18  # Catégorie
    ws.column_dimensions['B'].width = 18  # Module
    ws.column_dimensions['C'].width = 50  # Tâches
    ws.column_dimensions['D'].width = 15  # Type
    ws.column_dimensions['E'].width = 12  # Qui
    ws.column_dimensions['F'].width = 12  # Estimation
    
    # Fixer la hauteur de l'en-tête
    ws.row_dimensions[1].height = 30
    
    # Sauvegarder le fichier
    wb.save(output_file)
    print(f"✓ Fichier Excel créé: {output_file}")
    print(f"✓ Nombre de tâches: {len(tasks)}")

def main():
    """Fonction principale"""
    # Vérifier que le fichier markdown existe
    if not Path(MD_FILE).exists():
        print(f"✗ Erreur: Le fichier '{MD_FILE}' n'existe pas")
        return
    
    # Parser le markdown
    print(f"Lecture du fichier {MD_FILE}...")
    tasks = parse_markdown(MD_FILE)
    
    if not tasks:
        print("✗ Aucune tâche trouvée dans le fichier markdown")
        return
    
    print(f"✓ {len(tasks)} tâches trouvées")
    
    # Créer le fichier Excel
    print(f"Création du fichier Excel {EXCEL_FILE}...")
    create_excel(tasks, EXCEL_FILE)
    print("✓ Terminé!")

if __name__ == "__main__":
    main()
